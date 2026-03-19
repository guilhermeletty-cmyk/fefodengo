"""
Painel Logístico Dengo – v5.0
Módulos:
  1. Auditoria FEFO     – Quebras, devoluções ao Rack, plano de ação
  2. Análise de Pedidos – Crítica de múltiplos + Gatilho de Reabastecimento
     (Cenário A: OK/Sobra | Cenário B: Ruptura → múltiplo → Pallet do Rack)
"""

import streamlit as st
import pandas as pd
import numpy as np
import io, re, math, unicodedata
from datetime import datetime

# ══════════════════════════════════════════════════════════════════════════════
#  CONFIGURAÇÃO
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Operação Dengo – Painel Logístico",
    page_icon="📦",
    layout="wide"
)

st.markdown("""
<style>
[data-testid="stSidebar"]{background:#1a1a2e;}
[data-testid="stSidebar"] *{color:#e0e0e0 !important;}
.main-title{font-size:2rem;font-weight:700;color:#1a1a2e;}
.sub-title{font-size:1rem;color:#555;margin-bottom:1.2rem;}
.card{background:#f8f9fa;border-left:5px solid #0d6efd;padding:.9rem 1.1rem;border-radius:6px;margin-bottom:.6rem;}
.card-red{border-left-color:#dc3545;background:#fff5f5;}
.card-green{border-left-color:#198754;background:#f0fff4;}
.card-warn{border-left-color:#ffc107;background:#fffdf0;}
.card-blue{border-left-color:#0d6efd;background:#e8f4fd;}
.card-orange{border-left-color:#fd7e14;background:#fff3e0;}
.card-purple{border-left-color:#6f42c1;background:#f3f0ff;}
.metric-box{text-align:center;padding:1.1rem;border-radius:10px;box-shadow:0 2px 6px rgba(0,0,0,.08);}
.metric-title{font-size:.82rem;color:#666;margin-bottom:3px;}
.metric-value{font-size:1.9rem;font-weight:700;}
hr-sep{border:0;border-top:2px solid #e0e0e0;margin:1rem 0;}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## 📦 Operação Dengo")
    st.markdown("---")
    modulo = st.radio(
        "Módulo:",
        ["🔍 Auditoria FEFO", "📋 Análise de Pedidos"],
        label_visibility="collapsed"
    )
    st.markdown("---")
    st.caption("v5.0 | Painel Logístico Dengo")


# ══════════════════════════════════════════════════════════════════════════════
#  UTILITÁRIOS COMPARTILHADOS
# ══════════════════════════════════════════════════════════════════════════════

def limpar_sku(s: str) -> str:
    """Remove .0/.00 e espaços. Preserva alfanuméricos intactos."""
    s = str(s).strip()
    s = re.sub(r'\.0+$', '', s)
    return s


def _deacento(s: str) -> str:
    return ''.join(
        c for c in unicodedata.normalize('NFKD', str(s))
        if not unicodedata.combining(c)
    ).upper().strip()


def ler_arquivo(f) -> pd.DataFrame:
    nome = f.name.lower()
    raw  = f.read()
    if nome.endswith('.xlsx'):
        return pd.read_excel(io.BytesIO(raw))
    if nome.endswith('.xls'):
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=raw, ignore_workbook_corruption=True)
            ws = wb.sheet_by_index(0)
            data = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
            return pd.DataFrame(data[1:], columns=data[0])
        except Exception:
            pass
        return pd.read_excel(io.BytesIO(raw), engine='xlrd')
    for enc in ['utf-8', 'latin-1', 'cp1252']:
        for sep in [';', ',', '\t']:
            try:
                df = pd.read_csv(io.BytesIO(raw), sep=sep, encoding=enc)
                if df.shape[1] > 2:
                    return df
            except Exception:
                continue
    raise ValueError("Formato não suportado. Use CSV, XLS ou XLSX.")


def _safe_map(styled, fn, subset):
    """Compatível com pandas 1.x (applymap) e 2.x (map)."""
    try:
        return styled.map(fn, subset=subset)
    except AttributeError:
        return styled.applymap(fn, subset=subset)


def _excel_engine() -> str:
    """Detecta qual engine Excel está disponível no ambiente."""
    try:
        import xlsxwriter  # noqa
        return 'xlsxwriter'
    except ImportError:
        pass
    try:
        import openpyxl  # noqa
        return 'openpyxl'
    except ImportError:
        pass
    return 'none'


def _limpar_valor(val):
    """
    Converte qualquer valor para tipo seguro para escrita em Excel.
    Evita NaN, Inf, NaT e tipos numpy não serializáveis.
    """
    import math
    if val is None or val is pd.NaT:
        return ''
    if hasattr(val, 'item'):          # numpy scalar → Python nativo
        val = val.item()
    if isinstance(val, float):
        if math.isnan(val) or math.isinf(val):
            return ''
    return val


def gerar_excel(sheets: dict) -> bytes:
    """
    Gera um arquivo Excel (.xlsx) em memória com múltiplas abas coloridas.
    Regra xlsxwriter: NUNCA chamar df.to_excel() + ws.write() na mesma aba
    (causa dupla escrita / corrupção). Aqui escrevemos TUDO manualmente.
    Fallback automático para openpyxl ou ZIP-CSV se nenhum engine disponível.
    """
    STATUS_COLORS = {
        'INCORRETO': 'FFCCCC', 'QUEBRA':    'FFCCCC',
        'DEVOLVER':  'FFCCCC', 'RUPTURA':   'FFCCCC',
        'BUSCAR':    'CCE5FF', 'MOVER':     'CCE5FF',
        'REABASTEC': 'CCE5FF',
        'CORRETO':   'D4EDDA', 'OK':        'D4EDDA',
        'SOBRA':     'D4EDDA',
    }
    DEFAULT_COLOR = 'FFE0B2'
    HDR_BG        = '1a1a2e'

    engine = _excel_engine()
    buf    = io.BytesIO()

    # ── Fallback: ZIP com CSVs ─────────────────────────────────────────────
    if engine == 'none':
        import zipfile
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            for sname, (df, _) in sheets.items():
                if not isinstance(df, pd.DataFrame) or df.empty:
                    continue
                csv_bytes = df.to_csv(index=False, sep=';',
                                      encoding='utf-8-sig').encode('utf-8-sig')
                zf.writestr(f"{sname[:31]}.csv", csv_bytes)
        return buf.getvalue()

    # ── engine: xlsxwriter ────────────────────────────────────────────────
    # IMPORTANTE: não misturar df.to_excel() com ws.write() na mesma aba.
    # Escrevemos TUDO manualmente para evitar dupla escrita e corrupção.
    if engine == 'xlsxwriter':
        # nan_inf_to_errors=True converte NaN/Inf em erros Excel (não trava)
        writer_opts = {'options': {'nan_inf_to_errors': True,
                                   'strings_to_numbers': False}}
        with pd.ExcelWriter(buf, engine='xlsxwriter',
                            engine_kwargs=writer_opts) as writer:
            wb = writer.book

            fmt_hdr = wb.add_format({
                'bold': True, 'font_color': '#FFFFFF',
                'bg_color': f'#{HDR_BG}', 'align': 'center',
                'valign': 'vcenter', 'border': 1, 'text_wrap': True,
            })
            fmt_cache = {}
            def _fmt(color):
                if color not in fmt_cache:
                    fmt_cache[color] = wb.add_format(
                        {'bg_color': f'#{color}', 'border': 1,
                         'valign': 'vcenter'})
                return fmt_cache[color]

            for sname, (df, col_status) in sheets.items():
                if not isinstance(df, pd.DataFrame) or df.empty:
                    continue

                # Limpar todo o DataFrame antes de escrever
                df_clean = df.copy()
                for col in df_clean.columns:
                    df_clean[col] = df_clean[col].apply(_limpar_valor)

                sname31    = sname[:31]
                ws         = wb.add_worksheet(sname31)
                # Registrar a aba no writer para compatibilidade
                writer.sheets[sname31] = ws

                cols       = list(df_clean.columns)
                status_idx = cols.index(col_status) if (col_status and col_status in cols) else None

                # ── Linha de cabeçalho (linha 0) ──
                ws.set_row(0, 20)
                for ci, col in enumerate(cols):
                    ws.write(0, ci, str(col), fmt_hdr)

                # ── Linhas de dados (linha 1 em diante) ──
                for ri, row_vals in enumerate(df_clean.itertuples(index=False), start=1):
                    # Determinar cor da linha pelo valor de status
                    if status_idx is not None:
                        v     = str(row_vals[status_idx]).upper()
                        color = DEFAULT_COLOR
                        for kw, c in STATUS_COLORS.items():
                            if kw in v:
                                color = c
                                break
                    else:
                        color = DEFAULT_COLOR
                    fmt_row = _fmt(color)

                    for ci, val in enumerate(row_vals):
                        ws.write(ri, ci, val, fmt_row)

                # ── Largura automática das colunas ──
                for ci, col in enumerate(cols):
                    max_w = max(
                        len(str(col)),
                        df_clean[col].astype(str).str.len().max() if len(df_clean) > 0 else 0,
                    )
                    ws.set_column(ci, ci, min(int(max_w) + 4, 55))

    # ── engine: openpyxl (fallback) ───────────────────────────────────────
    else:
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils  import get_column_letter

        def _fill(hex6): return PatternFill('solid', fgColor=hex6)
        status_fills = {k: _fill(v) for k, v in STATUS_COLORS.items()}
        dflt_fill    = _fill(DEFAULT_COLOR)
        hdr_fill     = _fill(HDR_BG)
        hdr_font     = Font(color='FFFFFF', bold=True)

        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            for sname, (df, col_status) in sheets.items():
                if not isinstance(df, pd.DataFrame) or df.empty:
                    continue
                # Limpar NaN antes de escrever
                df_clean = df.fillna('').copy()
                df_clean.to_excel(writer, index=False, sheet_name=sname[:31])
                ws = writer.sheets[sname[:31]]
                for cell in ws[1]:
                    cell.fill = hdr_fill
                    cell.font = hdr_font
                    cell.alignment = Alignment(horizontal='center')
                if col_status and col_status in df_clean.columns:
                    idx = df_clean.columns.get_loc(col_status) + 1
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        v = str(row[idx - 1].value or '').upper()
                        fill = dflt_fill
                        for kw, f in status_fills.items():
                            if kw in v:
                                fill = f; break
                        for cell in row:
                            cell.fill = fill
                for ci in range(1, ws.max_column + 1):
                    ml = max((len(str(c.value or ''))
                              for row in ws.iter_rows(min_col=ci, max_col=ci)
                              for c in row), default=10)
                    ws.column_dimensions[get_column_letter(ci)].width = min(ml + 4, 55)

    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  MÓDULO 1 – AUDITORIA FEFO
# ══════════════════════════════════════════════════════════════════════════════

def classificar_setor(loc: str) -> str:
    try:
        rua   = str(loc)[2].upper()
        nivel = int(str(loc)[-1])
        if rua in ('L', 'M'): return 'Picking' if nivel in (1, 2) else 'Rack'
        if rua in ('J', 'K'): return 'Picking' if nivel == 1       else 'Rack'
    except Exception:
        pass
    return 'Rack'


def _normalizar_fefo(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [_deacento(c) for c in df.columns]
    alias = {
        'COD. PRODUTO':'COD_PRODUTO','COD PRODUTO':'COD_PRODUTO',
        'CODIGO PRODUTO':'COD_PRODUTO','SKU':'COD_PRODUTO',
        'DESCRICAO':'DESCRICAO','PRODUTO':'DESCRICAO','NOME':'DESCRICAO','ITEM':'DESCRICAO',
        'DATA DE VALIDADE':'DATA_VALIDADE','VALIDADE':'DATA_VALIDADE',
        'DATA VALIDADE':'DATA_VALIDADE','VENCIMENTO':'DATA_VALIDADE',
        'NR. LOTE':'LOTE','NR LOTE':'LOTE','LOTE':'LOTE','BATCH':'LOTE',
        'QTD UN':'QTD_UN','QTDE UN':'QTD_UN','QUANTIDADE':'QTD_UN','QTD':'QTD_UN','ESTOQUE':'QTD_UN',
        'LOCALIZACAO':'LOCALIZACAO','LOCAL':'LOCALIZACAO',
        'POSICAO':'LOCALIZACAO','ENDERECO':'LOCALIZACAO',
        'LOCALIZACAO ESTOQUE':'LOCALIZACAO',
    }
    df = df.rename(columns={k: v for k, v in alias.items() if k in df.columns})
    final_map = {}
    for col in df.columns:
        if col in alias.values(): continue
        for k, v in alias.items():
            if k in col and v not in df.columns and v not in final_map.values():
                final_map[col] = v; break
    return df.rename(columns=final_map)


def auditar_fefo(df: pd.DataFrame):
    df = _normalizar_fefo(df)
    ausentes = [c for c in ['COD_PRODUTO','DATA_VALIDADE','LOCALIZACAO'] if c not in df.columns]
    if ausentes:
        raise ValueError(f"Coluna(s) ausente(s): {ausentes} | Recebidas: {list(df.columns)}")
    df = df.dropna(subset=['COD_PRODUTO','DATA_VALIDADE','LOCALIZACAO']).copy()
    df['DATA_VALIDADE'] = pd.to_datetime(df['DATA_VALIDADE'], dayfirst=True, errors='coerce')
    df = df.dropna(subset=['DATA_VALIDADE'])
    df['COD_PRODUTO'] = df['COD_PRODUTO'].apply(limpar_sku)
    df['TIPO_LOC']    = df['LOCALIZACAO'].apply(classificar_setor)

    quebras, devolucao, skus_ok = [], [], []
    for sku, grp in df.groupby('COD_PRODUTO'):
        pick = grp[grp['TIPO_LOC'] == 'Picking']
        rack = grp[grp['TIPO_LOC'] == 'Rack']
        if pick.empty or rack.empty: skus_ok.append(sku); continue
        venc_pick_min = pick['DATA_VALIDADE'].min()
        criticos = rack[rack['DATA_VALIDADE'] < venc_pick_min].sort_values('DATA_VALIDADE')
        if criticos.empty: skus_ok.append(sku); continue
        venc_rack_min  = criticos['DATA_VALIDADE'].min()
        loc_pick_dest  = pick.sort_values('DATA_VALIDADE').iloc[0]['LOCALIZACAO']
        for _, l in criticos.iterrows():
            dias = (venc_pick_min - l['DATA_VALIDADE']).days
            descr = l.get('DESCRICAO', '—') if 'DESCRICAO' in l.index else '—'
            quebras.append({'SKU':sku,'DESCRICAO':descr,'ACAO':'BUSCAR DO RACK → PICKING',
                'LOTE':l.get('LOTE','—'),'LOCAL_ORIGEM':l['LOCALIZACAO'],
                'VENC_ORIGEM':l['DATA_VALIDADE'].strftime('%d/%m/%Y'),
                'LOCAL_DESTINO':loc_pick_dest,'VENC_NO_PICKING':venc_pick_min.strftime('%d/%m/%Y'),
                'DIAS_DIFERENCA':int(dias),'QTD':l.get('QTD_UN','—')})
        pick_errado = pick[pick['DATA_VALIDADE'] > venc_rack_min].sort_values('DATA_VALIDADE', ascending=False)
        for _, l in pick_errado.iterrows():
            descr = l.get('DESCRICAO','—') if 'DESCRICAO' in l.index else '—'
            devolucao.append({'SKU':sku,'DESCRICAO':descr,'ACAO':'DEVOLVER AO RACK',
                'LOTE':l.get('LOTE','—'),'LOCAL_PICKING':l['LOCALIZACAO'],
                'VENC_NO_PICKING':l['DATA_VALIDADE'].strftime('%d/%m/%Y'),
                'MOTIVO':f"Existe lote mais antigo no Rack (vence {venc_rack_min.strftime('%d/%m/%Y')})",
                'QTD':l.get('QTD_UN','—')})
    return pd.DataFrame(quebras), pd.DataFrame(devolucao), skus_ok


def render_fefo():
    st.markdown('<div class="main-title">🔍 Auditoria FEFO – Operação Dengo</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-title">Identifica quebras de FEFO, lotes a buscar no Rack e lotes a devolver do Picking.</div>', unsafe_allow_html=True)
    st.divider()

    with st.sidebar:
        st.markdown("**Regras de Picking**")
        st.info("Ruas L/M → Níveis 1 e 2\nRuas J/K → Nível 1")
        dias_critico = st.slider("Dias mínimos p/ alerta crítico", 0, 365, 30)

    uploaded = st.file_uploader("📂 Relatório de Estoque WMS (CSV / XLS / XLSX)", type=["csv","xlsx","xls"])
    if not uploaded:
        st.markdown('<div class="card"><b>Colunas esperadas:</b><br>'
                    '<code>LOCALIZACAO · COD. PRODUTO · DESCRICAO · DATA DE VALIDADE · NR. LOTE · QTD UN</code></div>',
                    unsafe_allow_html=True)
        return

    with st.spinner("Processando estoque..."):
        try:
            df_raw = ler_arquivo(uploaded)
            df_quebras, df_devol, skus_ok = auditar_fefo(df_raw)
            df_norm = _normalizar_fefo(df_raw)
            total_skus = df_norm['COD_PRODUTO'].nunique() if 'COD_PRODUTO' in df_norm.columns else '—'
            total_q  = len(df_quebras)
            total_d  = len(df_devol)
            criticas = int((df_quebras['DIAS_DIFERENCA'] >= dias_critico).sum()) if not df_quebras.empty else 0

            c1,c2,c3,c4,c5 = st.columns(5)
            for col,label,val,cor,bg in [
                (c1,"SKUs Analisados",  total_skus,"#0d6efd","#e8f4fd"),
                (c2,"SKUs Conformes",   len(skus_ok),"#198754","#f0fff4"),
                (c3,"Quebras FEFO",     total_q,"#dc3545" if total_q>0 else "#198754","#fff5f5" if total_q>0 else "#f0fff4"),
                (c4,f"Críticas ≥{dias_critico}d",criticas,"#dc3545" if criticas>0 else "#198754","#fff5f5" if criticas>0 else "#f0fff4"),
                (c5,"Devolver ao Rack", total_d,"#fd7e14" if total_d>0 else "#198754","#fff3e0" if total_d>0 else "#f0fff4"),
            ]:
                col.markdown(f'<div class="metric-box" style="background:{bg};"><div class="metric-title">{label}</div>'
                             f'<div class="metric-value" style="color:{cor};">{val}</div></div>', unsafe_allow_html=True)

            st.divider()
            if df_quebras.empty and df_devol.empty:
                st.success("✅ Picking 100% em conformidade com FEFO! Nenhuma movimentação necessária.")
                return

            plano = []
            for _, r in df_devol.iterrows():
                plano.append({'ORDEM':1,'ACAO':'1. DEVOLVER AO RACK','SKU':r['SKU'],'DESCRICAO':r['DESCRICAO'],
                    'LOTE':r['LOTE'],'DE':r['LOCAL_PICKING'],'PARA':'RACK (posição livre)','QTD':r['QTD'],'OBSERVACAO':r['MOTIVO']})
            for _, r in df_quebras.iterrows():
                plano.append({'ORDEM':2,'ACAO':'2. BUSCAR DO RACK','SKU':r['SKU'],'DESCRICAO':r['DESCRICAO'],
                    'LOTE':r['LOTE'],'DE':r['LOCAL_ORIGEM'],'PARA':r['LOCAL_DESTINO'],'QTD':r['QTD'],
                    'OBSERVACAO':f"Urgência: {r['DIAS_DIFERENCA']} dias antecipado"})
            df_plano = pd.DataFrame(plano).sort_values(['SKU','ORDEM']).reset_index(drop=True)

            def hl_dias(v):
                try:
                    vi = int(v)
                    if vi >= dias_critico: return 'background:#ffcccc;font-weight:bold;color:#c0392b'
                    if vi >= dias_critico//2: return 'background:#fff3cd;color:#856404'
                except: pass
                return ''

            def hl_acao(v):
                v = str(v)
                if '1.' in v or 'DEVOLVER' in v.upper(): return 'background:#ffe0b2;color:#7f3c00;font-weight:600'
                if '2.' in v or 'BUSCAR' in v.upper():   return 'background:#cce5ff;color:#004085;font-weight:600'
                return ''

            tab1,tab2,tab3 = st.tabs([
                f"🔴 Quebras FEFO – Buscar do Rack ({total_q})",
                f"🔁 Devolver ao Rack ({total_d})",
                f"📋 Plano de Ação ({len(df_plano)})",
            ])

            with tab1:
                st.markdown("**Lotes mais antigos parados no Rack → devem vir para o Picking:**")
                if df_quebras.empty:
                    st.success("Nenhuma quebra encontrada.")
                else:
                    for _, r in df_quebras.sort_values('DIAS_DIFERENCA', ascending=False).iterrows():
                        urg = "🔴 URGENTE" if r['DIAS_DIFERENCA'] >= dias_critico else "🟡 ATENÇÃO"
                        st.markdown(f'<div class="card card-red"><b>{r["DESCRICAO"]}</b> | SKU: <b>{r["SKU"]}</b> {urg} – <b>{r["DIAS_DIFERENCA"]} dias</b><br>'
                                    f'🏭 Origem Rack: {r["LOCAL_ORIGEM"]} – Lote {r["LOTE"]} – Vence {r["VENC_ORIGEM"]}<br>'
                                    f'📦 Destino Picking: {r["LOCAL_DESTINO"]} – Vence {r["VENC_NO_PICKING"]} | Qtd: {r["QTD"]}</div>',
                                    unsafe_allow_html=True)
                    st.dataframe(_safe_map(df_quebras.sort_values('DIAS_DIFERENCA',ascending=False).style, hl_dias, subset=['DIAS_DIFERENCA']),
                                 use_container_width=True, height=360)

            with tab2:
                st.markdown("**Lotes no Picking mais novos → devem voltar ao Rack:**")
                if df_devol.empty:
                    st.success("Nenhum item a devolver.")
                else:
                    for _, r in df_devol.iterrows():
                        st.markdown(f'<div class="card card-orange"><b>{r["DESCRICAO"]}</b> | SKU: <b>{r["SKU"]}</b><br>'
                                    f'📦 Picking: {r["LOCAL_PICKING"]} – Lote {r["LOTE"]} – Vence {r["VENC_NO_PICKING"]}<br>'
                                    f'⚠️ {r["MOTIVO"]} | Qtd: {r["QTD"]}</div>', unsafe_allow_html=True)
                    st.dataframe(_safe_map(df_devol.style, hl_acao, subset=['ACAO']), use_container_width=True, height=360)

            with tab3:
                st.markdown("**Sequência recomendada: 1. Devolver → 2. Buscar (mesmo SKU):**")
                if not df_plano.empty:
                    st.dataframe(_safe_map(df_plano.style, hl_acao, subset=['ACAO']), use_container_width=True, height=480)

            st.divider()
            ts = datetime.now().strftime('%Y%m%d_%H%M')
            c1,c2 = st.columns(2)
            with c1:
                xb = gerar_excel({"Quebras FEFO–Buscar":(df_quebras,'ACAO'),"Devolver ao Rack":(df_devol,'ACAO'),"Plano de Acao":(df_plano,'ACAO')})
                st.download_button("📊 Excel Completo FEFO", xb, f"FEFO_{ts}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            with c2:
                st.download_button("📄 CSV – Quebras",
                    df_quebras.to_csv(index=False,sep=';',encoding='utf-8-sig').encode('utf-8-sig'),
                    f"FEFO_QUEBRAS_{ts}.csv","text/csv", use_container_width=True)

        except Exception as e:
            st.error(f"❌ Erro: {e}"); st.exception(e)


# ══════════════════════════════════════════════════════════════════════════════
#  MÓDULO 2 – ANÁLISE DE PEDIDOS  (v5: Gatilho de Reabastecimento)
# ══════════════════════════════════════════════════════════════════════════════

def _estoque_picking_por_sku(df_estoque_raw: pd.DataFrame) -> pd.DataFrame:
    """
    A partir do relatório de estoque WMS, retorna:
      SKU | ESTOQUE_PICKING | linhas de Rack com: SKU, LOCAL, LOTE, VALIDADE, QTD_UN
    """
    est = _normalizar_fefo(df_estoque_raw)
    est['COD_PRODUTO'] = est['COD_PRODUTO'].apply(limpar_sku)
    est['QTD_UN']      = pd.to_numeric(est.get('QTD_UN', pd.Series(dtype=float)),
                                        errors='coerce').fillna(0)
    est['TIPO_LOC']    = est['LOCALIZACAO'].apply(classificar_setor)

    picking = (
        est[est['TIPO_LOC'] == 'Picking']
        .groupby('COD_PRODUTO', as_index=False)['QTD_UN']
        .sum()
        .rename(columns={'COD_PRODUTO': 'SKU', 'QTD_UN': 'ESTOQUE_PICKING'})
    )

    rack = est[est['TIPO_LOC'] == 'Rack'].copy()
    # Ordenar por validade para respeitar FEFO na sugestão de pallet
    if 'DATA_VALIDADE' in rack.columns:
        rack['DATA_VALIDADE'] = pd.to_datetime(rack['DATA_VALIDADE'], dayfirst=True, errors='coerce')
        rack = rack.sort_values(['COD_PRODUTO', 'DATA_VALIDADE'])

    return picking, rack


def calcular_gatilho_reabastecimento(
    df_demanda: pd.DataFrame,          # SKU | PRODUTO | DEMANDA_TOTAL | MULTIPLO_UNCX
    df_picking: pd.DataFrame,          # SKU | ESTOQUE_PICKING
    df_rack: pd.DataFrame,             # COD_PRODUTO | LOCALIZACAO | LOTE | DATA_VALIDADE | QTD_UN
) -> tuple:
    """
    Retorna (df_gatilho, df_pallets_mover)

    df_gatilho: linha por SKU com
        SKU | PRODUTO | DEMANDA | ESTOQUE_PICKING | SALDO_PICKING | STATUS_GATILHO
        | NECESSIDADE_ABAST | NECESSIDADE_ARREDONDADA | CX_NECESSARIAS

    df_pallets_mover: linha por pallet sugerido para Cenário B
        SKU | PRODUTO | PALLET_ID | LOTE | VALIDADE | QTD_PALLET | UN_MOVER | CX_MOVER | ACAO
    """
    merged = df_demanda.merge(df_picking, on='SKU', how='left')
    merged['ESTOQUE_PICKING'] = pd.to_numeric(
        merged['ESTOQUE_PICKING'], errors='coerce'
    ).fillna(0)
    merged['MULTIPLO_UNCX'] = pd.to_numeric(
        merged['MULTIPLO_UNCX'], errors='coerce'
    ).fillna(1).replace(0, 1)

    dem    = merged['DEMANDA_TOTAL'].values.astype(float)
    pick   = merged['ESTOQUE_PICKING'].values.astype(float)
    mult   = merged['MULTIPLO_UNCX'].values.astype(float)

    saldo  = pick - dem                         # positivo = sobra, negativo = falta
    nec    = np.where(saldo < 0, -saldo, 0.0)   # necessidade bruta (unidades faltando)

    # Arredondar para cima ao múltiplo da caixa
    nec_arr = np.where(
        nec > 0,
        np.ceil(nec / mult) * mult,
        0.0
    )
    cx_nec = np.where(mult > 0, np.ceil(nec_arr / mult), 0).astype(int)

    status = np.where(
        saldo >= 0,
        'OK - SEM NECESSIDADE DE REABASTECIMENTO',
        'RUPTURA - REABASTECIMENTO NECESSÁRIO'
    )

    df_gatilho = pd.DataFrame({
        'SKU':                    merged['SKU'].values,
        'PRODUTO':                merged['PRODUTO'].values,
        'DEMANDA_TOTAL':          dem.astype(int),
        'ESTOQUE_PICKING':        pick.astype(int),
        'SALDO_PICKING':          saldo.astype(int),
        'STATUS_GATILHO':         status,
        'NECESSIDADE_BRUTA_UN':   nec.astype(int),
        'NECESSIDADE_ARREDON_UN': nec_arr.astype(int),
        'CX_A_MOVER':             cx_nec,
        'MULTIPLO_UNCX':          mult.astype(int),
    })

    # ── Sugestão de Pallets (somente Cenário B) ──────────────────────
    rupturas = df_gatilho[df_gatilho['STATUS_GATILHO'].str.startswith('RUPTURA')].copy()
    pallets_linhas = []

    for _, row in rupturas.iterrows():
        sku         = row['SKU']
        cx_demanda  = int(row['CX_A_MOVER'])
        mult_sku    = float(row['MULTIPLO_UNCX'])
        rack_sku    = df_rack[df_rack['COD_PRODUTO'] == sku].copy()

        acumulado_cx = 0
        for _, pal in rack_sku.iterrows():
            if acumulado_cx >= cx_demanda:
                break
            qtd_pal  = float(pal['QTD_UN'])
            cx_disp  = math.floor(qtd_pal / mult_sku) if mult_sku > 0 else 0
            cx_usar  = min(cx_disp, cx_demanda - acumulado_cx)
            if cx_usar <= 0:
                continue
            un_usar = int(cx_usar * mult_sku)

            venc_str = ''
            if 'DATA_VALIDADE' in pal.index:
                try:
                    venc_str = pd.Timestamp(pal['DATA_VALIDADE']).strftime('%d/%m/%Y')
                except Exception:
                    pass

            pallets_linhas.append({
                'SKU':          sku,
                'PRODUTO':      row['PRODUTO'],
                'PALLET_ID':    pal['LOCALIZACAO'],
                'LOTE':         pal.get('LOTE', '—'),
                'VALIDADE':     venc_str,
                'QTD_NO_PALLET': int(qtd_pal),
                'CX_DISPONIVEIS': int(cx_disp),
                'CX_MOVER':     int(cx_usar),
                'UN_MOVER':     un_usar,
                'MULTIPLO_UNCX': int(mult_sku),
                'ACAO': (f"Mover {un_usar} un ({cx_usar} cx)"
                         f" do Pallet {pal['LOCALIZACAO']} → PICKING"),
            })
            acumulado_cx += cx_usar

    df_pallets = pd.DataFrame(pallets_linhas)
    return df_gatilho, df_pallets


def analisar_pedidos(
    df_wms_raw: pd.DataFrame,
    df_ped_raw: pd.DataFrame,
    df_estoque_raw=None,
):
    # ── WMS Cadastro ──────────────────────────────────────────────────
    wms = df_wms_raw.copy()
    wms.columns = wms.columns.str.strip().str.upper()
    sku_col  = wms.columns[1]          # B = COD. PRODUTO
    nome_col = wms.columns[2]          # C = PRODUTO
    mult_col = wms.columns[7]          # H = UN/CX  (SEMPRE índice 7)
    pal_col  = wms.columns[10] if wms.shape[1] > 10 else None  # K = CX/PALETE

    cols = [sku_col, nome_col, mult_col] + ([pal_col] if pal_col else [])
    wms_c = wms[cols].copy()
    wms_c.columns = ['SKU_WMS', 'PRODUTO', 'MULTIPLO'] + (['CX_PALETE'] if pal_col else [])
    wms_c['SKU_WMS']  = wms_c['SKU_WMS'].apply(limpar_sku)
    wms_c['MULTIPLO'] = pd.to_numeric(wms_c['MULTIPLO'], errors='coerce').fillna(1).replace(0, 1)
    if 'CX_PALETE' in wms_c.columns:
        wms_c['CX_PALETE'] = pd.to_numeric(wms_c['CX_PALETE'], errors='coerce').fillna(0)
    else:
        wms_c['CX_PALETE'] = 0.0

    # ── Pedidos ───────────────────────────────────────────────────────
    ped = df_ped_raw.copy()
    ped.columns = ped.columns.str.strip().str.upper()
    sku_col_ped = next((c for c in ped.columns if 'COD' in c and ('PROD' in c or 'SKU' in c)), ped.columns[7])
    qtd_col     = next((c for c in ped.columns if 'QTDE' in c or ('QTD' in c and 'UN' not in c)), ped.columns[9])
    nf_col      = ped.columns[0]

    ped_c = ped[[nf_col, sku_col_ped, qtd_col]].copy()
    ped_c.columns = ['PEDIDO_NF', 'SKU_PED', 'QTDE_REQ']
    ped_c['SKU_PED']  = ped_c['SKU_PED'].apply(limpar_sku)
    ped_c['QTDE_REQ'] = pd.to_numeric(ped_c['QTDE_REQ'], errors='coerce').fillna(0)
    ped_c = ped_c[ped_c['QTDE_REQ'] > 0].reset_index(drop=True)

    # ── Merge pedidos × cadastro ──────────────────────────────────────
    merged = ped_c.merge(wms_c, left_on='SKU_PED', right_on='SKU_WMS', how='left')
    merged['MULTIPLO']  = pd.to_numeric(merged['MULTIPLO'],  errors='coerce').fillna(1).replace(0, 1)
    merged['PRODUTO']   = merged['PRODUTO'].fillna('NAO CADASTRADO NO WMS')
    merged['CX_PALETE'] = pd.to_numeric(merged.get('CX_PALETE', pd.Series(dtype=float)),
                                         errors='coerce').fillna(0)

    # ── Cálculos crítica de múltiplo ──────────────────────────────────
    qtd_v    = merged['QTDE_REQ'].values.astype(float)
    mult_v   = merged['MULTIPLO'].values.astype(float)
    cx_pal_v = merged['CX_PALETE'].values.astype(float)
    divisao_v = np.round(qtd_v / mult_v, 4)
    ideal_v   = np.ceil(qtd_v / mult_v) * mult_v
    diferenca_v = (ideal_v - qtd_v).astype(int)
    cx_nec_v  = np.ceil(ideal_v / mult_v).astype(int)
    paletes_v = np.where(cx_pal_v > 0, np.ceil(cx_nec_v / cx_pal_v), 0).astype(int)
    status_v  = np.where(np.abs(qtd_v % mult_v) < 1e-6, 'CORRETO', 'INCORRETO')

    df_geral = pd.DataFrame({
        'PEDIDO_NF':      merged['PEDIDO_NF'].values,
        'SKU':            merged['SKU_PED'].values,
        'PRODUTO':        merged['PRODUTO'].values,
        'QTDE_REQUERIDA': qtd_v,
        'MULTIPLO_UNCX':  mult_v,
        'DIVISAO':        divisao_v,
        'STATUS':         status_v,
        'QTDE_IDEAL':     ideal_v,
        'DIFERENCA_UN':   diferenca_v,
        'CX_NECESSARIAS': cx_nec_v,
        'CX_POR_PALETE':  cx_pal_v.astype(int),
        'PALETES_ESTIM':  paletes_v,
    })

    df_erros = df_geral[df_geral['STATUS'] == 'INCORRETO'].copy().reset_index(drop=True)
    if not df_erros.empty:
        df_erros['SUGESTAO'] = (
            'Ajustar de ' + df_erros['QTDE_REQUERIDA'].astype(int).astype(str)
            + ' para '    + df_erros['QTDE_IDEAL'].astype(int).astype(str)
            + ' (+'       + df_erros['DIFERENCA_UN'].astype(int).astype(str) + ' un)'
        )

    df_reabast = (
        df_geral.groupby(['SKU','PRODUTO','MULTIPLO_UNCX'], as_index=False).agg(
            TOTAL_UN  =('QTDE_IDEAL',     'sum'),
            TOTAL_CX  =('CX_NECESSARIAS', 'sum'),
            CX_PALETE =('CX_POR_PALETE',  'first'),
            PALETES   =('PALETES_ESTIM',  'sum'),
            STATUS    =('STATUS', lambda x: 'INCORRETO' if 'INCORRETO' in x.values else 'CORRETO'),
        ).sort_values('TOTAL_UN', ascending=False).reset_index(drop=True)
    )

    # ── Gatilho de Reabastecimento (Cenário A / B) ────────────────────
    df_gatilho   = pd.DataFrame()
    df_pallets   = pd.DataFrame()
    if df_estoque_raw is not None:
        try:
            df_picking_est, df_rack_est = _estoque_picking_por_sku(df_estoque_raw)

            # Consolidar demanda por SKU (usando QTDE_IDEAL já múltipla)
            demanda_sku = (
                df_geral.groupby(['SKU','PRODUTO','MULTIPLO_UNCX'], as_index=False)
                .agg(DEMANDA_TOTAL=('QTDE_IDEAL','sum'))
            )

            df_gatilho, df_pallets = calcular_gatilho_reabastecimento(
                demanda_sku.rename(columns={'MULTIPLO_UNCX':'MULTIPLO_UNCX'}),
                df_picking_est,
                df_rack_est,
            )
        except Exception as ex:
            df_gatilho = pd.DataFrame({'ERRO': [str(ex)]})

    return df_geral, df_erros, df_reabast, df_gatilho, df_pallets


# ── Colunas exatas do layout WMS Dengo (ordem obrigatória para reimportação) ──
_COLUNAS_WMS_SAIDA = [
    'OE / VIAGEM',
    'CNPJ EMPRESA',
    'PLACA',
    'CNPJ TRANSPORTADORA',
    'CPF MOTORISTA',
    'CLIENTE / FORNECEDOR',
    'PEDIDOS / NF',          # coluna G – NF sequencial gerada pelo sistema
    'COD PRODUTO',
    'LOTE',
    'QTDE REQUERIDA',        # preenchida com QTDE_IDEAL (já múltipla)
    'PESO LIQ.',
    'PESO BRUTO',
    'DATA AGENDAMENTO (DD/MM/YYYY)',
    'TIPO EXPEDIÇÃO (1 - normal / 2 - cross)',
    'Valor NF',
]


def _montar_layout_wms(
    df_ped_raw: pd.DataFrame,
    df_geral: pd.DataFrame,
    nf_inicio: int = 15215,
) -> pd.DataFrame:
    """
    Reconstrói o DataFrame no layout exato do arquivo de pedidos WMS,
    substituindo:
      • coluna G (PEDIDOS / NF) → NF sequencial única, sem repetição,
        partindo de `nf_inicio`.
      • coluna J (QTDE REQUERIDA) → QTDE_IDEAL (quantidade corrigida ao múltiplo).
    Preserva todas as demais colunas originais.
    """
    # Normalizar colunas do pedido original
    ped = df_ped_raw.copy()
    ped.columns = ped.columns.str.strip()

    # Detectar colunas chave do arquivo original
    def _find_col(df, *keywords):
        for kw in keywords:
            matches = [c for c in df.columns if kw.upper() in str(c).upper()]
            if matches:
                return matches[0]
        return None

    sku_col  = _find_col(ped, 'COD PRODUTO', 'COD. PRODUTO', 'SKU')
    qtd_col  = _find_col(ped, 'QTDE REQUERIDA', 'QTDE', 'QTD')
    nf_col   = _find_col(ped, 'PEDIDOS / NF', 'PEDIDOS/NF', 'NF')

    # Limpar SKU no arquivo original para cruzar com df_geral
    if sku_col:
        ped['_SKU_CLEAN'] = ped[sku_col].apply(limpar_sku)
    else:
        ped['_SKU_CLEAN'] = ''

    # Preservar ARQUIVO_ORIGEM (coluna interna para análise – não vai ao WMS)
    arquivo_origem_series = ped['ARQUIVO_ORIGEM'].copy() if 'ARQUIVO_ORIGEM' in ped.columns else None
    if 'ARQUIVO_ORIGEM' in ped.columns:
        ped = ped.drop(columns=['ARQUIVO_ORIGEM'])

    # Filtrar apenas linhas com QTDE > 0 (mesmo filtro do analisar_pedidos)
    if qtd_col:
        ped[qtd_col] = pd.to_numeric(ped[qtd_col], errors='coerce').fillna(0)
        ped = ped[ped[qtd_col] > 0].reset_index(drop=True)

    # Garantir que df_geral e ped têm o mesmo tamanho após filtro
    n = min(len(ped), len(df_geral))
    ped    = ped.iloc[:n].copy().reset_index(drop=True)
    geral  = df_geral.iloc[:n].copy().reset_index(drop=True)

    # ── Gerar NF por PEDIDO (não por linha) ─────────────────────────────
    # Uma NF única para cada pedido (OE / VIAGEM); linhas do mesmo pedido
    # compartilham a mesma NF.
    viagem_col = _find_col(ped, 'OE / VIAGEM', 'OE/VIAGEM', 'OE VIAGEM', 'VIAGEM', 'OE')
    if viagem_col and viagem_col in ped.columns:
        # Mapear cada pedido distinto a um número NF sequencial
        pedidos_unicos = list(dict.fromkeys(ped[viagem_col].tolist()))  # preserva ordem
        nf_map         = {p: nf_inicio + i for i, p in enumerate(pedidos_unicos)}
        nfs            = ped[viagem_col].map(nf_map).tolist()
        ultimo_nf_calc = nf_inicio + len(pedidos_unicos) - 1
    else:
        # Fallback: uma NF por linha (comportamento anterior)
        nfs            = list(range(nf_inicio, nf_inicio + n))
        ultimo_nf_calc = nfs[-1] if nfs else nf_inicio

    # ── Substituir colunas ────────────────────────────────────────────────
    if nf_col and nf_col in ped.columns:
        ped[nf_col] = nfs
    elif 'PEDIDOS / NF' in ped.columns:
        ped['PEDIDOS / NF'] = nfs
    else:
        # Inserir na posição G (índice 6) se não existir
        ped.insert(min(6, len(ped.columns)), 'PEDIDOS / NF', nfs)

    # Substituir QTDE REQUERIDA pela quantidade ideal (já múltipla)
    qtde_ideal_col = 'QTDE_IDEAL' if 'QTDE_IDEAL' in geral.columns else None
    if qtd_col and qtd_col in ped.columns and qtde_ideal_col:
        ped[qtd_col] = geral[qtde_ideal_col].values.astype(int)

    # ── Reordenar para layout WMS ──────────────────────────────────────────
    # Mapear nomes reais das colunas para os nomes padrão WMS
    col_rename = {}
    for std_name in _COLUNAS_WMS_SAIDA:
        match = _find_col(ped, std_name)
        if match and match != std_name:
            col_rename[match] = std_name
    ped = ped.rename(columns=col_rename)

    # Selecionar somente colunas do layout, criando as ausentes como vazio
    for c in _COLUNAS_WMS_SAIDA:
        if c not in ped.columns:
            ped[c] = ''
    saida = ped[_COLUNAS_WMS_SAIDA].copy()

    # ── Reincorporar ARQUIVO_ORIGEM como última coluna (só análise) ────────
    if arquivo_origem_series is not None:
        # Realinhar índice após os filtros/resets anteriores
        saida = saida.reset_index(drop=True)
        ao = arquivo_origem_series.reset_index(drop=True).iloc[:len(saida)]
        saida['ARQUIVO_ORIGEM'] = ao.values

    return saida, ultimo_nf_calc   # retorna df + último NF usado


def _ler_multiplos_pedidos(lista_arquivos) -> pd.DataFrame:
    """
    Lê uma lista de arquivos de pedidos (CSV/XLS/XLSX) e concatena
    em um único DataFrame, adicionando a coluna ARQUIVO_ORIGEM para rastreio.
    """
    frames = []
    erros  = []
    for f in lista_arquivos:
        try:
            df = ler_arquivo(f)
            df['ARQUIVO_ORIGEM'] = f.name
            frames.append(df)
        except Exception as ex:
            erros.append(f"{f.name}: {ex}")
    if erros:
        st.warning("⚠️ Arquivos ignorados por erro de leitura:\n" + "\n".join(f"• {e}" for e in erros))
    if not frames:
        raise ValueError("Nenhum arquivo de pedidos foi lido com sucesso.")
    return pd.concat(frames, ignore_index=True)


def render_pedidos():
    st.markdown('<div class="main-title">📋 Análise de Pedidos – Crítica de Múltiplos</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="sub-title">'
        'Cruza cadastro WMS (col. H = UN/CX) com demanda. '
        'Aceita múltiplos arquivos de pedidos. '
        'Gera relatório no layout do WMS com NF sequencial configurável.'
        '</div>', unsafe_allow_html=True
    )
    st.divider()

    # ── Configuração de NF na sidebar ─────────────────────────────────────
    with st.sidebar:
        st.markdown("---")
        st.markdown("**⚙️ Configuração de NF**")
        nf_inicio = st.number_input(
            "Número inicial da NF",
            min_value=1,
            max_value=9_999_999,
            value=15215,
            step=1,
            help=(
                "O sistema gerará NFs sequenciais únicas a partir deste número. "
                "Altere em caso de recuperação de dados ou reinício de sequência. "
                "Padrão: 15215."
            )
        )
        st.caption(f"Próxima NF gerada: **{int(nf_inicio)}**")

    col1, col2, col3 = st.columns(3)
    with col1:
        f_wms = st.file_uploader(
            "📦 Cadastro WMS",
            type=["csv","xlsx","xls"],
            help="Col B=SKU | Col H=UN/CX (múltiplo) | Col K=CX/PALETE"
        )
    with col2:
        f_peds = st.file_uploader(
            "🛒 Pedidos do Cliente (1 ou + arquivos)",
            type=["csv","xlsx","xls"],
            accept_multiple_files=True,
            help="Selecione quantos arquivos quiser (30+). Todos serão consolidados automaticamente."
        )
    with col3:
        f_est = st.file_uploader(
            "🏭 Estoque WMS (obrigatório para Gatilho)",
            type=["csv","xlsx","xls"],
            help="Mesmo relatório do módulo FEFO: LOCALIZACAO, COD. PRODUTO, QTD UN, DATA VALIDADE, LOTE"
        )

    if not f_wms or not f_peds:
        st.markdown("""<div class="card">
        <b>Instruções:</b><br>
        1️⃣ <b>Cadastro WMS</b> – col H = UN/CX (múltiplo), col K = CX/PALETE<br>
        2️⃣ <b>Pedidos do Cliente</b> – selecione <b>1 ou mais arquivos de uma vez</b> (sem limite); todos são consolidados<br>
        3️⃣ <b>Estoque WMS</b> (opcional) – habilita o <b>Gatilho de Reabastecimento</b>
             (compara estoque picking com demanda e indica pallets do Rack)
        </div>""", unsafe_allow_html=True)
        return

    # ── Painel de resumo dos arquivos de pedidos carregados ────────────
    with st.expander(f"📂 {len(f_peds)} arquivo(s) de pedidos carregado(s) – clique para ver detalhes",
                     expanded=(len(f_peds) <= 5)):
        col_a, col_b = st.columns([3, 1])
        with col_a:
            for f in f_peds:
                kb = round(f.size / 1024, 1)
                st.markdown(f"• `{f.name}` &nbsp;&nbsp;({kb} KB)", unsafe_allow_html=True)
        with col_b:
            st.metric("Total de arquivos", len(f_peds))

    with st.spinner(f"Lendo {len(f_peds)} arquivo(s) de pedidos e cruzando dados..."):
        try:
            df_wms = ler_arquivo(f_wms)
            df_ped = _ler_multiplos_pedidos(f_peds)   # ← consolida todos os pedidos
            df_est = ler_arquivo(f_est) if f_est else None

            # Informar quantas linhas foram consolidadas
            st.info(f"📋 Pedidos consolidados: **{len(df_ped):,} linhas** de **{len(f_peds)} arquivo(s)**")

            df_geral, df_erros, df_reabast, df_gatilho, df_pallets = analisar_pedidos(
                df_wms, df_ped, df_est
            )

            # ── Montar layout WMS com NF sequencial ────────────────────
            df_wms_saida, ultimo_nf = _montar_layout_wms(
                df_ped, df_geral, nf_inicio=int(nf_inicio)
            )

            total_itens = len(df_geral)
            total_erros = len(df_erros)
            total_ok    = total_itens - total_erros
            pct_ok      = round(total_ok / total_itens * 100, 1) if total_itens > 0 else 0

            # KPIs linha 1 – crítica de múltiplos
            total_cx = int(df_reabast['TOTAL_CX'].sum()) if 'TOTAL_CX' in df_reabast.columns else 0
            total_un = int(df_reabast['TOTAL_UN'].sum()) if 'TOTAL_UN' in df_reabast.columns else 0
            total_pl = int(df_reabast['PALETES'].sum())  if 'PALETES'  in df_reabast.columns else 0

            c1,c2,c3,c4,c5 = st.columns(5)
            for col,label,val,cor,bg in [
                (c1,"Itens no Pedido",   total_itens,"#0d6efd","#e8f4fd"),
                (c2,"Múltiplo Correto",  total_ok,"#198754","#f0fff4"),
                (c3,"Erros de Múltiplo", total_erros,
                    "#dc3545" if total_erros>0 else "#198754",
                    "#fff5f5" if total_erros>0 else "#f0fff4"),
                (c4,"Taxa de Acerto",    f"{pct_ok}%",
                    "#198754" if pct_ok>=90 else "#ffc107" if pct_ok>=50 else "#dc3545",
                    "#f0fff4" if pct_ok>=90 else "#fffdf0" if pct_ok>=50 else "#fff5f5"),
                (c5,"~Paletes Demanda",  total_pl,"#6f42c1","#f3f0ff"),
            ]:
                col.markdown(f'<div class="metric-box" style="background:{bg};"><div class="metric-title">{label}</div>'
                             f'<div class="metric-value" style="color:{cor};">{val}</div></div>', unsafe_allow_html=True)

            # KPIs linha 2 – gatilho (somente se estoque foi enviado)
            if not df_gatilho.empty and 'ERRO' not in df_gatilho.columns and 'STATUS_GATILHO' in df_gatilho.columns:
                st.markdown("---")
                n_ok      = int((df_gatilho['STATUS_GATILHO'].str.startswith('OK')).sum())
                n_rupt    = int((df_gatilho['STATUS_GATILHO'].str.startswith('RUPTURA')).sum())
                un_falt   = int(df_gatilho['NECESSIDADE_ARREDON_UN'].sum())
                cx_falt   = int(df_gatilho['CX_A_MOVER'].sum())
                pallets_n = len(df_pallets)

                ca,cb,cc,cd,ce = st.columns(5)
                for col,label,val,cor,bg in [
                    (ca,"✅ SKUs OK (Picking suficiente)", n_ok, "#198754","#f0fff4"),
                    (cb,"🚨 SKUs em Ruptura",              n_rupt,"#dc3545" if n_rupt>0 else "#198754",
                        "#fff5f5" if n_rupt>0 else "#f0fff4"),
                    (cc,"📦 UN a Reabastecer",              un_falt,"#fd7e14" if un_falt>0 else "#198754",
                        "#fff3e0" if un_falt>0 else "#f0fff4"),
                    (cd,"📫 Caixas a Mover",                cx_falt,"#0d6efd","#e8f4fd"),
                    (ce,"🏭 Pallets Sugeridos",             pallets_n,"#6f42c1","#f3f0ff"),
                ]:
                    col.markdown(f'<div class="metric-box" style="background:{bg};"><div class="metric-title">{label}</div>'
                                 f'<div class="metric-value" style="color:{cor};">{val}</div></div>', unsafe_allow_html=True)

            st.divider()

            # ── Funções de cor ─────────────────────────────────────────
            def hl_status_multi(v):
                v = str(v).upper()
                if 'INCORRETO' in v: return 'background:#f8d7da;color:#721c24;font-weight:600'
                if 'CORRETO'   in v: return 'background:#d4edda;color:#155724;font-weight:600'
                return ''

            def hl_gatilho(v):
                v = str(v).upper()
                if 'RUPTURA'    in v: return 'background:#f8d7da;color:#721c24;font-weight:700'
                if 'OK'         in v: return 'background:#d4edda;color:#155724;font-weight:600'
                return ''

            def hl_pallet(v):
                if 'Mover' in str(v): return 'background:#cce5ff;color:#004085;font-weight:600'
                return ''

            # ── Construir lista de tabs dinamicamente ─────────────────
            abas = [
                f"📄 Relatório Geral ({total_itens})",
                f"🚨 Painel de Erros ({total_erros})",
                f"📊 Gatilho de Reabastecimento",
                f"🏭 Pallets a Movimentar",
                f"📦 Reabastecimento Consolidado",
            ]
            tabs = st.tabs(abas)

            # TAB 0 – Relatório Geral (layout WMS)
            with tabs[0]:
                # Contar pedidos únicos (NFs distintas geradas)
                n_pedidos_unicos = df_wms_saida['PEDIDOS / NF'].nunique() \
                    if 'PEDIDOS / NF' in df_wms_saida.columns else '?'
                n_arquivos_unicos = df_wms_saida['ARQUIVO_ORIGEM'].nunique() \
                    if 'ARQUIVO_ORIGEM' in df_wms_saida.columns else '?'

                st.markdown(
                    f"**Layout pronto para importação no WMS** – "
                    f"**{n_pedidos_unicos} pedido(s) distintos** → NF **{int(nf_inicio)}** "
                    f"até **{ultimo_nf}** | {n_arquivos_unicos} arquivo(s) de origem:"
                )
                st.caption(
                    "✅ Coluna G (PEDIDOS / NF) – uma NF por pedido (OE / VIAGEM), linhas do mesmo pedido compartilham a mesma NF · "
                    "✅ QTDE REQUERIDA corrigida pelo múltiplo · "
                    "✅ Colunas na ordem exata do WMS · "
                    "ℹ️ Coluna ARQUIVO_ORIGEM = apenas para análise (não enviada ao WMS)"
                )
                # Highlight: NF em azul, ARQUIVO_ORIGEM em cinza
                def hl_nf(v):
                    try:
                        if int(v) >= int(nf_inicio):
                            return 'background:#e8f4fd;color:#004085;font-weight:600'
                    except Exception:
                        pass
                    return ''

                def hl_arquivo(v):
                    return 'background:#f8f9fa;color:#495057;font-style:italic'

                col_nf  = 'PEDIDOS / NF'
                col_arq = 'ARQUIVO_ORIGEM'
                style_df = df_wms_saida.style
                if col_nf  in df_wms_saida.columns:
                    style_df = _safe_map(style_df, hl_nf,     subset=[col_nf])
                if col_arq in df_wms_saida.columns:
                    style_df = _safe_map(style_df, hl_arquivo, subset=[col_arq])
                st.dataframe(style_df, use_container_width=True, height=480)

            # TAB 1 – Painel de Erros
            with tabs[1]:
                if df_erros.empty:
                    st.success("✅ Nenhum erro de múltiplo encontrado! Pedido 100% correto.")
                else:
                    st.warning(f"⚠️ {total_erros} item(ns) com quantidade não múltipla de UN/CX:")
                    for _, r in df_erros.iterrows():
                        st.markdown(f"""<div class="card card-red">
                            <b>{r['PRODUTO']}</b> | SKU: {r['SKU']} | Pedido: {r['PEDIDO_NF']}<br>
                            Múltiplo (UN/CX): <b>{int(r['MULTIPLO_UNCX'])}</b> |
                            Qtde pedida: <b>{int(r['QTDE_REQUERIDA'])}</b>
                            ÷ {int(r['MULTIPLO_UNCX'])} = <b>{r['DIVISAO']}</b> (fracionado)<br>
                            ✅ Sugestão: {r.get('SUGESTAO','—')}
                        </div>""", unsafe_allow_html=True)
                    st.dataframe(_safe_map(df_erros.style, hl_status_multi, subset=['STATUS']),
                                 use_container_width=True, height=300)

            # TAB 2 – Gatilho de Reabastecimento
            with tabs[2]:
                st.markdown("### 🔁 Gatilho de Reabastecimento – Estoque Picking × Demanda")
                st.caption("Cenário A ✅ – Estoque Picking ≥ Demanda → sem movimentação | "
                           "Cenário B 🚨 – Estoque Picking < Demanda → reabastecimento do Rack")

                if df_gatilho.empty:
                    st.info("📂 Envie o arquivo de **Estoque WMS** para habilitar esta análise.")
                elif 'ERRO' in df_gatilho.columns:
                    st.error(f"Erro ao processar estoque: {df_gatilho['ERRO'].iloc[0]}")
                elif 'STATUS_GATILHO' not in df_gatilho.columns:
                    st.warning("Dados insuficientes para calcular gatilho.")
                else:
                    rupturas_df = df_gatilho[df_gatilho['STATUS_GATILHO'].str.startswith('RUPTURA')]
                    ok_df       = df_gatilho[df_gatilho['STATUS_GATILHO'].str.startswith('OK')]

                    if not rupturas_df.empty:
                        st.error(f"🚨 {len(rupturas_df)} SKU(s) em RUPTURA – picking insuficiente para atender a demanda:")
                        for _, r in rupturas_df.sort_values('NECESSIDADE_ARREDON_UN', ascending=False).iterrows():
                            st.markdown(f"""<div class="card card-red">
                                <b>{r['PRODUTO']}</b> | SKU: <b>{r['SKU']}</b><br>
                                📊 Demanda: <b>{int(r['DEMANDA_TOTAL'])} un</b> |
                                📦 Estoque Picking: <b>{int(r['ESTOQUE_PICKING'])} un</b> |
                                ❌ Falta: <b>{int(r['NECESSIDADE_BRUTA_UN'])} un</b><br>
                                ➡️ Reabastecimento (arredondado ao múltiplo {int(r['MULTIPLO_UNCX'])}):
                                <b>{int(r['NECESSIDADE_ARREDON_UN'])} un / {int(r['CX_A_MOVER'])} cx</b>
                            </div>""", unsafe_allow_html=True)

                    if not ok_df.empty:
                        with st.expander(f"✅ {len(ok_df)} SKU(s) com Estoque Picking suficiente (Cenário A)"):
                            for _, r in ok_df.iterrows():
                                sobra = int(r['SALDO_PICKING'])
                                st.markdown(f"""<div class="card card-green">
                                    <b>{r['PRODUTO']}</b> | SKU: <b>{r['SKU']}</b><br>
                                    📊 Demanda: {int(r['DEMANDA_TOTAL'])} un |
                                    📦 Estoque Picking: {int(r['ESTOQUE_PICKING'])} un |
                                    ✅ Sobra: <b>{sobra} un</b>
                                </div>""", unsafe_allow_html=True)

                    st.markdown("#### 📋 Tabela Completa – Gatilho por SKU")
                    st.dataframe(
                        _safe_map(df_gatilho.style, hl_gatilho, subset=['STATUS_GATILHO']),
                        use_container_width=True, height=420
                    )

            # TAB 3 – Pallets a Movimentar
            with tabs[3]:
                st.markdown("### 🏭 Pallets do Rack indicados para Reabastecimento do Picking")
                st.caption("Ordenados por FEFO (validade mais próxima primeiro). "
                           "Apenas SKUs em Cenário B (Ruptura).")

                if df_pallets.empty:
                    if df_gatilho.empty:
                        st.info("📂 Envie o arquivo de **Estoque WMS** para habilitar esta análise.")
                    else:
                        st.success("✅ Nenhum pallet precisa ser movimentado – estoque picking suficiente!")
                elif 'ERRO' in df_pallets.columns:
                    st.warning(str(df_pallets['ERRO'].iloc[0]))
                else:
                    # Agrupar por SKU para exibição visual
                    for sku, grp in df_pallets.groupby('SKU'):
                        produto = grp['PRODUTO'].iloc[0]
                        total_un = int(grp['UN_MOVER'].sum())
                        total_cx = int(grp['CX_MOVER'].sum())
                        n_pal    = len(grp)

                        with st.expander(f"📦 SKU {sku} – {produto} | {total_un} un / {total_cx} cx / {n_pal} pallet(s)"):
                            for _, r in grp.iterrows():
                                st.markdown(f"""<div class="card card-blue">
                                    🏭 <b>Pallet ID:</b> {r['PALLET_ID']}
                                    | Lote: {r['LOTE']}
                                    | Validade: {r['VALIDADE']}<br>
                                    📊 Estoque no Pallet: {r['QTD_NO_PALLET']} un ({r['CX_DISPONIVEIS']} cx disponíveis)<br>
                                    ➡️ <b>Mover: {r['UN_MOVER']} un ({r['CX_MOVER']} cx)</b> → PICKING
                                </div>""", unsafe_allow_html=True)

                    st.markdown("#### 📋 Tabela Completa – Pallets a Movimentar")
                    st.dataframe(
                        _safe_map(df_pallets.style, hl_pallet, subset=['ACAO']),
                        use_container_width=True, height=420
                    )

            # TAB 4 – Reabastecimento Consolidado
            with tabs[4]:
                st.markdown("**Consolidado por SKU – caixas e paletes estimados (base demanda):**")
                st.dataframe(_safe_map(df_reabast.style, hl_status_multi, subset=['STATUS']),
                             use_container_width=True, height=480)
                st.info(f"📦 Total demanda: **{total_un} un** | **{total_cx} cx** | **~{total_pl} paletes**")

            # ── Downloads ──────────────────────────────────────────────
            st.divider()
            ts = datetime.now().strftime('%Y%m%d_%H%M')

            # Versão WMS pura (sem ARQUIVO_ORIGEM) – pronta para reimportar
            df_wms_puro = df_wms_saida.drop(columns=['ARQUIVO_ORIGEM'], errors='ignore')

            cd1, cd2 = st.columns(2)
            with cd1:
                sheets_dict = {
                    # Aba "Geral"  = layout WMS PURO (sem ARQUIVO_ORIGEM) → importar no WMS
                    "Geral":           (df_wms_puro,  'PEDIDOS / NF'),
                    # Aba "Geral+Arquivo" = com coluna de rastreio de origem
                    "Geral+Arquivo":   (df_wms_saida, 'PEDIDOS / NF'),
                    "Erros Multiplo":  (df_erros,     'STATUS'),
                    "Reabast Consol":  (df_reabast,   'STATUS'),
                }
                if not df_gatilho.empty and 'STATUS_GATILHO' in df_gatilho.columns:
                    sheets_dict["Gatilho Reabast"] = (df_gatilho, 'STATUS_GATILHO')
                if not df_pallets.empty and 'ACAO' in df_pallets.columns:
                    sheets_dict["Pallets a Mover"] = (df_pallets, 'ACAO')
                xb = gerar_excel(sheets_dict)
                st.download_button(
                    f"📊 Excel WMS – NF {int(nf_inicio)} a {ultimo_nf}",
                    xb, f"PEDIDOS_WMS_NF{int(nf_inicio)}_{ts}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                st.caption("Aba **Geral** = WMS puro · Aba **Geral+Arquivo** = com rastreio de origem do arquivo")
            with cd2:
                if not df_gatilho.empty and 'STATUS_GATILHO' in df_gatilho.columns:
                    rupt = df_gatilho[df_gatilho['STATUS_GATILHO'].str.startswith('RUPTURA')]
                    csv_d = rupt.to_csv(index=False,sep=';',encoding='utf-8-sig').encode('utf-8-sig')
                    st.download_button("📄 CSV – Somente Rupturas", csv_d, f"RUPTURAS_{ts}.csv",
                        "text/csv", use_container_width=True)
                else:
                    csv_d = df_erros.to_csv(index=False,sep=';',encoding='utf-8-sig').encode('utf-8-sig') if not df_erros.empty else df_geral.to_csv(index=False,sep=';',encoding='utf-8-sig').encode('utf-8-sig')
                    st.download_button("📄 CSV – Erros / Geral", csv_d, f"PEDIDOS_GERAL_{ts}.csv",
                        "text/csv", use_container_width=True)

        except Exception as e:
            st.error(f"❌ Erro: {e}"); st.exception(e)


# ══════════════════════════════════════════════════════════════════════════════
#  ROTEADOR
# ══════════════════════════════════════════════════════════════════════════════
if modulo == "🔍 Auditoria FEFO":
    render_fefo()
else:
    render_pedidos()
